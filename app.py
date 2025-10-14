# full_streamlit_shift_planner.py
import streamlit as st
import pandas as pd
import numpy as np
from calendar import monthrange, weekday
import io  # for Excel export

st.set_page_config(layout="wide", page_title="Employee Leave & Shift Planner")
st.title("Employee Leave & Shift Planner — 24/7 Coverage")

# --------------------------
# 1. Employee / Group Definition
# --------------------------
employee_data = pd.DataFrame([
    ["Gopalakrishnan Selvaraj", "IIS", "Any"],      # PrefShift: what shifts they can/ prefer (e.g., "F-S", "Any")
    ["Paneerselvam F", "IIS", "Any"],
    ["Rajesh Jayapalan", "IIS", "Any"],
    ["Ajay Chidipotu", "Websphere", "Any"],
    ["Imran Khan", "Websphere", "F-S"],            # Cannot do N
    ["Sammeta Balachander", "Websphere", "Any"],
    ["Ramesh Polisetty", "Middleware", "Any"],
    ["Srinivasu Cheedalla", "Middleware", "Any"],
    ["Gangavarapu Suneetha", "Middleware", "Any"],
    ["Lakshmi Narayana Rao", "Middleware", "Any"],
    ["Muppa Divya", "Middleware", "Any"],
    ["Anil Athkuri", "Middleware", "Any"],
    ["D Namithananda", "Middleware", "Any"],
    ["Pousali C", "", "Any"],
    ["Thorat Yashwant", "", "Any"],
    ["Srivastav Nitin", "", "Any"],
    ["Kishore Khati Vaibhav", "", "Any"],
    ["Rupan Venkatesan Anandha", "", "Any"],
    ["Chaudhari Kaustubh", "", "Any"],
    ["Shejal Gawade", "", "Any"],
    ["Vivek Kushwaha", "", "Any"],
    ["Abdul Mukthiyar Basha", "", "Any"],
    ["M Naveen", "", "Any"],
    ["B Madhurusha", "", "Any"],
    ["Chinthalapudi Yaswanth", "", "Any"],
    ["Edagotti Kalpana", "", "Any"]
], columns=["Name", "Skill", "PrefShift"])

employees = employee_data["Name"].tolist()

# Groups
group1 = ["Gopalakrishnan Selvaraj", "Paneerselvam F", "Rajesh Jayapalan"]
group2 = ["Ajay Chidipotu", "Imran Khan", "Sammeta Balachander"]
group3 = ["Ramesh Polisetty", "Srinivasu Cheedalla", "Gangavarapu Suneetha", "Lakshmi Narayana Rao"]
group4 = ["Muppa Divya", "Anil Athkuri", "D Namithananda"]
coverage_pool = [
    "Pousali C", "B Madhurusha", "Chinthalapudi Yaswanth", "Edagotti Kalpana",
    "Thorat Yashwant", "Srivastav Nitin", "Kishore Khati Vaibhav",
    "Rupan Venkatesan Anandha", "Chaudhari Kaustubh", "Shejal Gawade", "Vivek Kushwaha"
]

# --------------------------
# 2. UI Tabs
# --------------------------
tab1, tab2, tab3, tab4 = st.tabs(["Config & Generate", "Individual Leave", "Preview / Edit", "Final Plan & Summary"])

# --------------------------
# 3. Helper functions
# --------------------------
def get_weekdays(year, month, weekday_indices):
    return [d for d in range(1, monthrange(year, month)[1] + 1) if weekday(year, month, d) in weekday_indices]

def get_weekoff_days_for_pattern(year, month, pattern):
    patterns = {
        "Friday-Saturday": [4, 5], "Sunday-Monday": [6, 0], "Saturday-Sunday": [5, 6],
        "Monday-Tuesday": [0, 1], "Tuesday-Wednesday": [1, 2],
        "Wednesday-Thursday": [2, 3], "Thursday-Friday": [3, 4]
    }
    idxs = patterns.get(pattern, [])
    return [d for d in range(1, monthrange(year, month)[1] + 1) if weekday(year, month, d) in idxs]

# -------------------------------------------------
# TAB 1 – Config & Generate
# -------------------------------------------------
with tab1:
    st.header("Step 1 – Month, Week-offs & Targets")
    col1, col2 = st.columns(2)
    with col1:
        year = st.number_input("Year", 2023, 2100, 2025)
        month = st.selectbox("Month", range(1,13), index=9,
                             format_func=lambda x: pd.Timestamp(year,x,1).strftime("%B"))
    with col2:
        seed = st.number_input("Random seed (re-producibility)", value=42)

    num_days = monthrange(year, month)[1]
    dates = [pd.Timestamp(year, month, d).strftime("%d-%b-%Y") for d in range(1, num_days+1)]

    # Week-off patterns
    weekoff_options = ["None", "Friday-Saturday", "Sunday-Monday", "Saturday-Sunday",
                       "Monday-Tuesday", "Tuesday-Wednesday", "Wednesday-Thursday", "Thursday-Friday"]
    weekoff_init = pd.DataFrame({emp: "None" for emp in employees}, index=weekoff_options[1:])
    weekoff_df = st.data_editor(weekoff_init, height=300)

    # Festivals
    festival_days = st.multiselect("Festival / Common holidays (dates 1-N)", range(1, num_days+1))

    # Targets
    col_target1, col_target2 = st.columns(2)
    with col_target1:
        target_morning = st.number_input("Target Morning (F)", 1, 15, 3)
        target_second  = st.number_input("Target Second (S)", 1, 15, 3)
    with col_target2:
        target_mid     = st.number_input("Target Mid (M)", 0, 15, 0)
        max_night_per_day = st.number_input("Max Night (N) per day", 1, 10, 2)
        max_nights_per_person = st.number_input("Max nights per person", 0, num_days, 6)

    st.info("**Coverage pool** (auto-filled from blanks): " + ", ".join(coverage_pool))

    if st.button("Generate Plan"):
        rng = np.random.default_rng(seed)

        # Build weekoff_dict
        weekoff_dict = {emp: "None" for emp in employees}
        for emp in employees:
            chosen = weekoff_df[emp].replace("None", None).dropna()
            if not chosen.empty:
                weekoff_dict[emp] = chosen.index[0]

        # Initialize plan
        plan = {emp: [''] * num_days for emp in employees}

        # 1. Apply week-offs
        for emp, pattern in weekoff_dict.items():
            if pattern == "None":
                continue
            off_days = get_weekoff_days_for_pattern(year, month, pattern)
            for d in off_days:
                plan[emp][d-1] = 'O'

        # 2. Apply festivals
        for emp in employees:
            for f in festival_days:
                plan[emp][f-1] = 'H'

        # 3. Fixed groups
        fixed_g3 = {
            "Ramesh Polisetty": "G", "Srinivasu Cheedalla": "E",
            "Gangavarapu Suneetha": "G", "Lakshmi Narayana Rao": "G"
        }
        for emp, sh in fixed_g3.items():
            for d in range(num_days):
                if plan[emp][d] == '':
                    plan[emp][d] = sh

        for emp in group4:
            for d in range(num_days):
                if plan[emp][d] == '':
                    plan[emp][d] = 'S'

        # 4. Group1 rotation
        cycle_g1 = ['F', 'S', 'N']
        for offset, emp in enumerate(group1):
            idx = offset
            cur = cycle_g1[idx % 3]
            for d in range(num_days):
                if plan[emp][d] in ('O', 'H'):
                    idx = (idx + 1) % 3
                    cur = cycle_g1[idx % 3]
                elif plan[emp][d] == '':
                    plan[emp][d] = cur

        # 5. Group2 rotation
        for offset, emp in enumerate(group2):
            if emp == "Imran Khan":
                cycle = ['F', 'S']
                idx = offset % 2
                cur = cycle[idx]
                for d in range(num_days):
                    if plan[emp][d] in ('O', 'H'):
                        idx = (idx + 1) % 2
                        cur = cycle[idx]
                    elif plan[emp][d] == '':
                        plan[emp][d] = cur
            else:
                idx = offset
                cur = cycle_g1[idx % 3]
                for d in range(num_days):
                    if plan[emp][d] in ('O', 'H'):
                        idx = (idx + 1) % 3
                        cur = cycle_g1[idx % 3]
                    elif plan[emp][d] == '':
                        plan[emp][d] = cur

        # Shift counts for fairness
        shift_count = {emp: {'F':0, 'S':0, 'M':0, 'N':0} for emp in employees}
        for emp in employees:
            for val in plan[emp]:
                if val in shift_count[emp]:
                    shift_count[emp][val] += 1

        # Preferences map
        pref_map = dict(zip(employee_data["Name"], employee_data["PrefShift"]))

        # Helper count
        def count_shift(day_idx, code):
            return sum(1 for e in employees if plan[e][day_idx] == code)

        # Greedy assignment per day
        for d in range(num_days):
            cur = {'F': count_shift(d, 'F'), 'S': count_shift(d, 'S'),
                   'M': count_shift(d, 'M'), 'N': count_shift(d, 'N')}

            need = {
                'F': max(0, target_morning - cur['F']),
                'S': max(0, target_second - cur['S']),
                'M': max(0, target_mid - cur['M']),
                'N': max(0, max_night_per_day - cur['N'])
            }

            candidates = [e for e in employees if plan[e][d] == '']
            rng.shuffle(candidates)

            # Assign N first
            for _ in range(need['N']):
                night_cands = [e for e in candidates
                               if shift_count[e]['N'] < max_nights_per_person
                               and 'N' in pref_map.get(e, "Any")]
                if not night_cands:
                    break
                chosen = min(night_cands, key=lambda x: (shift_count[x]['N'], rng.random()))
                plan[chosen][d] = 'N'
                shift_count[chosen]['N'] += 1
                candidates.remove(chosen)

            # Assign M, F, S
            for shift, tgt in [('M', need['M']), ('F', need['F']), ('S', need['S'])]:
                for _ in range(tgt):
                    if not candidates:
                        break
                    chosen = candidates.pop(0)
                    plan[chosen][d] = shift
                    shift_count[chosen][shift] += 1

            # Fill remaining
            remaining = [e for e in employees if plan[e][d] == '']
            for e in remaining:
                for sh, tgt in [('F', target_morning), ('S', target_second), ('M', target_mid)]:
                    if cur[sh] < tgt:
                        plan[e][d] = sh
                        shift_count[e][sh] += 1
                        cur[sh] += 1
                        break
                else:
                    plan[e][d] = 'S'
                    shift_count[e]['S'] += 1

        # Store
        df_plan = pd.DataFrame(plan, index=dates).T
        st.session_state['plan_raw'] = df_plan
        st.session_state['shift_count'] = pd.DataFrame(shift_count).T
        st.session_state['year'] = year
        st.session_state['month'] = month
        st.session_state['max_nights_per_person'] = max_nights_per_person
        st.success("Initial plan generated! Go to Preview / Edit to tweak.")

# -------------------------------------------------
# TAB 2 – Individual Leave
# -------------------------------------------------
with tab2:
    st.header("Individual leave / ad-hoc adjustments")
    if 'plan_raw' not in st.session_state:
        st.warning("Generate a plan first.")
    else:
        df = st.session_state['plan_raw'].copy()
        emp = st.selectbox("Employee", employees, key="leave_emp")
        days = st.multiselect("Date(s) for Leave (L)", df.columns, key="leave_days")
        if st.button("Apply Leave"):
            for d in days:
                df.loc[emp, d] = 'L'
            st.session_state['plan_raw'] = df
            st.success(f"Leave applied for {emp} on {', '.join(days)}")

# -------------------------------------------------
# TAB 3 – Preview & Manual Edit
# -------------------------------------------------
with tab3:
    st.header("Preview & Fine-tune")
    if 'plan_raw' not in st.session_state:
        st.warning("No plan yet.")
    else:
        df_edit = st.session_state['plan_raw'].copy()
        edited = st.data_editor(df_edit, height=600)
        if st.button("Lock this version"):
            st.session_state['final_plan'] = edited
            st.success("Plan locked! Go to Final Plan & Summary.")

# -------------------------------------------------
# TAB 4 – Final Plan & Summary
# -------------------------------------------------
with tab4:
    st.header("Final Shift Plan")
    if 'final_plan' not in st.session_state:
        st.warning("Lock a plan in the Preview tab first.")
    else:
        df_plan = st.session_state['final_plan']
        year = st.session_state.get('year', 2025)
        month = st.session_state.get('month', 1)
        max_nights_per_person = st.session_state.get('max_nights_per_person', 6)

        def color_map(val):
            cmap = {
                'F': '#d4f6d4', 'S': '#c3e0ff', 'M': '#ffff99', 'N': '#ffccdd',
                'G': '#ffd700', 'E': '#ee82ee', 'O': '#e6e6e6', 'H': '#ffb84d', 'L': '#ff6666'
            }
            return f'background-color: {cmap.get(val, "")}'
        styled = df_plan.style.applymap(color_map)
        st.dataframe(styled, height=600)

        # Totals
        totals = pd.DataFrame({
            sh: [sum(1 for v in df_plan.loc[e] if v == sh) for e in df_plan.index]
            for sh in ('F', 'S', 'M', 'N', 'G', 'E', 'O', 'H', 'L')
        }, index=df_plan.index)
        st.subheader("Shift totals per employee")
        st.dataframe(totals)

        # Daily
        daily = pd.DataFrame({
            'Date': df_plan.columns,
            'F': [sum(1 for e in df_plan.index if df_plan.loc[e, col] == 'F') for col in df_plan.columns],
            'S': [sum(1 for e in df_plan.index if df_plan.loc[e, col] == 'S') for col in df_plan.columns],
            'M': [sum(1 for e in df_plan.index if df_plan.loc[e, col] == 'M') for col in df_plan.columns],
            'N': [sum(1 for e in df_plan.index if df_plan.loc[e, col] == 'N') for col in df_plan.columns],
            'Off': [sum(1 for e in df_plan.index if df_plan.loc[e, col] in ('O', 'H', 'L')) for col in df_plan.columns]
        })
        st.subheader("Daily coverage")
        st.dataframe(daily)

        # Downloads - CSV always
        csv = df_plan.to_csv().encode('utf-8')
        st.download_button("Download CSV", csv, f"shift_plan_{year}_{month:02d}.csv", "text/csv")

        # Excel export with colors (uses pandas + openpyxl via engine)
        try:
            xlsx_bytes = io.BytesIO()
            with pd.ExcelWriter(xlsx_bytes, engine='openpyxl') as writer:
                df_plan.to_excel(writer, sheet_name='Plan')
                workbook = writer.book
                worksheet = writer.sheets['Plan']
                cmap = {
                    'F': 'd4f6d4', 'S': 'c3e0ff', 'M': 'ffff99', 'N': 'ffccdd',
                    'G': 'ffd700', 'E': 'ee82ee', 'O': 'e6e6e6', 'H': 'ffb84d', 'L': 'ff6666'
                }
                from openpyxl.styles import PatternFill
                for row_idx, emp in enumerate(df_plan.index, start=2):  # headers at row 1-2
                    for col_idx, date in enumerate(df_plan.columns, start=2):
                        val = df_plan.loc[emp, date]
                        if val in cmap:
                            fill = PatternFill(start_color=cmap[val], end_color=cmap[val], fill_type='solid')
                            worksheet.cell(row=row_idx, column=col_idx).fill = fill
            xlsx_bytes.seek(0)
            st.download_button("Download Excel (with colors)", xlsx_bytes.read(), f"shift_plan_{year}_{month:02d}.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.warning(f"Excel export failed ({str(e)}); falling back to CSV only. Ensure 'openpyxl' is in requirements.txt")

        # Compliance
        nights = totals['N']
        over = nights[nights > max_nights_per_person]
        if not over.empty:
            st.error(f"Warning: {len(over)} employee(s) exceed max-nights: {', '.join(over.index)}")
        else:
            st.success("Night limits OK!")
